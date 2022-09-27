import openpyxl as xl
from openpyxl.styles import PatternFill
from config import OUTPUT_DATA_INFO, REAL_DATA_INFO, CHECK_DATA_INFO, OUTPUT_SIMILARITY, WARN_COLOR, ORIGIN_COLOR, ERROR_COLOR


class ExcelWorker:
    def __init__(self, filename):
        self.filename = filename
        self.excel_en = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        self.excel_en_dict = {el: i + 1 for i, el in enumerate(self.excel_en)}
        self.data = None
        self.data_list = None

    def read_data(self):
        self.data = xl.load_workbook(self.filename)

    def read_list(self, list_name):
        if self.data is not None:
            self.data_list = self.data[list_name]

    def get_data(self, column, row_start, row_end):
        data = [self.data_list.cell(row=i, column=column).value for i in range(row_start+1, row_end)]
        return data

    def convert_data_info(self):
        check_info = (self.__convert_to_number(CHECK_DATA_INFO[0]), CHECK_DATA_INFO[1],
                      CHECK_DATA_INFO[2], CHECK_DATA_INFO[3])
        real_info = (self.__convert_to_number(REAL_DATA_INFO[0]), REAL_DATA_INFO[1],
                     REAL_DATA_INFO[2], REAL_DATA_INFO[3])

        return check_info, real_info

    def __convert_to_number(self, column):
        column = sum([self.excel_en_dict[el]*pow(len(self.excel_en), len(column)-i-1) for i, el in enumerate(column)])
        return column + 1

    def save_excel_data(self, data: dict, filename):
        warnFill = PatternFill(start_color=WARN_COLOR,
                               end_color=WARN_COLOR,
                               fill_type="solid")
        originFill = PatternFill(start_color=ORIGIN_COLOR,
                                 end_color=ORIGIN_COLOR,
                                 fill_type="solid")
        errorFill = PatternFill(start_color=ERROR_COLOR,
                                end_color=ERROR_COLOR,
                                fill_type="solid")
        column = self.__convert_to_number(column=OUTPUT_DATA_INFO[0]) - 1
        list_name = OUTPUT_DATA_INFO[1]
        row_start = OUTPUT_DATA_INFO[2]
        row_end = OUTPUT_DATA_INFO[3]

        items = list(data.items())
        if self.filename == filename:
            self.read_list(list_name=list_name)
        for i in range(row_start, row_end):
            cell = self.data_list.cell(row=i, column=column)
            next_cell = self.data_list.cell(row=i, column=column+1)
            key = items[i-row_start][0].split(" ")
            value = " ".join(items[i-row_start][-1][0])
            name, value = value.split("similarity")
            similarity = items[i-row_start][-1][1]
            cell.fill = errorFill if similarity < OUTPUT_SIMILARITY or len(name.split(",")) > 1 else originFill
            if len(key) > len(name.split(" ")):
                cell.fill = warnFill
            cell.value = name
            next_cell.value = value
        self.data.save(filename)
